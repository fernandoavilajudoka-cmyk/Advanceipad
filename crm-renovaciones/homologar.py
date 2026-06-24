#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Homologación de la base de clientes / renovaciones (Advance · Mapon CRM)
=======================================================================

Toma los 4 archivos de Excel históricos/operativos y produce UNA sola base
homologada por VIN, lista para sincronizarse contra el CRM de Mapon.

Llave de cruce de TODA la información: VIN.

Salida (carpeta ./salida):
  - base_homologada.xlsx   (hojas: Unidades, Empresas, Conflictos, Metodologia)
  - clientes_unidades.csv  (1 fila por VIN, encabezados alineados al CRM de Mapon)
  - empresas.csv           (1 fila por empresa, con tipo AAA/AA/A)

Regla de decisión para FECHA DE RENOVACIÓN cuando hay fechas distintas
(definida con el cliente):
  Prioridad  1) Operaciones  (RENOVACIONES ACTUALIZADO, validada a mano)
             2) Operaciones complementarias (Unidades nuevas / Hoja 34 / Missa / Agosto)
             3) Export BASE DE DATOS JUNIO (snapshot de la API)
             4) Bot_API (jalado crudo de la API)
  El CRM en vivo es la verdad final (ver reconciliar_crm.py).
  Si dos fuentes difieren > TOLERANCIA_DIAS la unidad se marca en la hoja
  "Conflictos" para revisión humana.

Clasificación de cliente (según número de unidades de la empresa):
  AAA = 20 unidades en adelante   ·  Prioridad Alta
  AA  = 10 a 19 unidades          ·  Prioridad Media
  A   = 1 a 9 unidades            ·  Prioridad Estándar

Uso:
  python3 homologar.py            # usa ./fuentes y ./salida
  python3 homologar.py --fuentes <dir> --salida <dir>
"""
import argparse
import csv
import datetime as dt
import glob
import os
import re
import sys
import unicodedata
from collections import defaultdict

try:
    import openpyxl
except ImportError:
    sys.exit("Falta openpyxl: pip install openpyxl")

TOLERANCIA_DIAS = 7  # diferencia máx. tolerada entre fuentes antes de marcar conflicto

# ----------------------------------------------------------------------------
# Helpers
# ----------------------------------------------------------------------------
def fold(s):
    """minúsculas, sin acentos, espacios colapsados."""
    if s is None:
        return ""
    s = str(s)
    s = unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode("ascii")
    s = re.sub(r"\s+", " ", s).strip().lower()
    return s

_BASURA = {"", "none", "#ref!", "pendiente", "n/a", "na", "sin dato", "-", "."}

def norm_vin(v):
    if v is None:
        return None
    s = str(v).strip().upper()
    s = re.sub(r"\s+", "", s)
    # algunas matrículas vienen como GN-<VIN>; quitamos el prefijo
    if s.startswith("GN-"):
        s = s[3:]
    if s.lower() in _BASURA:
        return None
    return s

def is_vin(s):
    return bool(s) and 11 <= len(s) <= 20 and re.fullmatch(r"[A-Z0-9]+", s) is not None

def norm_dispositivo(s):
    """De 'CAN006/GV350CEU' o 'ESENCIAL/GV58LAU' saca el modelo GPS (GV350CEU…)."""
    if not s:
        return ""
    u = str(s).upper()
    m = re.search(r"GV\d{2,3}[A-Z]*", u) or re.search(r"FM[A-Z]?\d{3,4}", u) or re.search(r"AT\d{2,3}", u)
    if m:
        return m.group(0)
    tok = u.split("/")[-1].strip()
    return tok if re.search(r"[A-Z]", tok) and len(tok) <= 14 else ""

def clean(v):
    if v is None:
        return ""
    s = str(v).strip()
    if s.lower() in _BASURA:
        return ""
    # teléfonos exportados como '8115318020.0'
    if re.fullmatch(r"\d+\.0", s):
        s = s[:-2]
    return s

def parse_date(v):
    """Devuelve datetime.date o None."""
    if v is None:
        return None
    if isinstance(v, dt.datetime):
        return v.date()
    if isinstance(v, dt.date):
        return v
    s = str(v).strip()
    if s.lower() in _BASURA:
        return None
    s = s.split(" ")[0]  # quita la hora
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%m/%d/%Y", "%d.%m.%Y", "%Y/%m/%d"):
        try:
            d = dt.datetime.strptime(s, fmt).date()
            if 2019 <= d.year <= 2035:   # descarta basura (años 0207, 1935, 2926…)
                return d
            return None
        except ValueError:
            continue
    return None

# ----------------------------------------------------------------------------
# Detección de columnas por encabezado (robusta a layouts distintos)
# ----------------------------------------------------------------------------
def classify(h):
    h0 = fold(h)
    if not h0:
        return None
    if h0 in ("vin", "vin number") or "num. de serie" in h0 or "num de serie" in h0 \
            or "numero de serie" in h0 or "num. de serie" in h0:
        return "vin"
    if "nombre de contacto" in h0 or "nombre del contacto" in h0:
        return "contacto_nombre"
    if h0 == "contacto":
        return "contacto_alt"
    if h0 in ("cliente", "clientes", "empresa", "compania", "nombre empresa",
              "nombre del cliente", "nombre de cuenta") or "nombre empresa" in h0:
        return "empresa"
    if "renov" in h0 or "vencimiento" in h0 or "venicimiento" in h0 or "venici" in h0:
        return "fecha_renov"
    if "fecha de alta" == h0 or ("alta" in h0 and "fecha" in h0) or "instalacion" in h0 \
            or "fecha de creacion" in h0:
        return "fecha_alta"
    if "telefono" in h0 or h0 == "tel" or "phone" in h0:
        return "telefono"
    if "correo" in h0 or "email" in h0 or "e-mail" in h0:
        return "correo"
    if "unidades" in h0:
        return "unidades"
    if "clasificacion" in h0 or "tipo de cliente" in h0 or h0 == "segmento":
        return "clasificacion"
    if "modelo gps" in h0:
        return "modelo_gps"
    if h0 == "dispositivo":          # hoja mensual: "ESENCIAL/GV58LAU", "CAN006/GV350CEU"
        return "dispositivo"
    if "estatus" in h0:
        return "estatus"
    if h0 == "vigencia":
        return "vigencia"
    return None

def map_headers(row):
    """row -> {canonico: idx}. Si hay choque, gana el primero salvo contacto_nombre."""
    m = {}
    for idx, cell in enumerate(row):
        c = classify(cell)
        if c and c not in m:
            m[c] = idx
    return m

def find_header(ws, max_scan=8):
    """Devuelve (idx_fila_0based, mapa). Elige la fila con más encabezados reconocidos."""
    best = (-1, {}, -1)
    rows = list(ws.iter_rows(min_row=1, max_row=min(max_scan, ws.max_row), values_only=True))
    for i, r in enumerate(rows):
        m = map_headers(r)
        score = len(m) + (2 if "vin" in m else 0) + (1 if "empresa" in m else 0)
        if "vin" in m and ("fecha_renov" in m or "fecha_alta" in m or "empresa" in m):
            if score > best[2]:
                best = (i, m, score)
    return best[0], best[1]

# ----------------------------------------------------------------------------
# Definición de fuentes: (archivo_glob, hoja, prioridad_renov, etiqueta)
# prioridad_renov menor = más confiable.  None = la hoja no aporta fecha de renovación.
# ----------------------------------------------------------------------------
SRC = "fuentes"  # se sobreescribe por args

def wb_of(nombre_contiene):
    hits = [f for f in glob.glob(os.path.join(SRC, "*.xlsx")) if nombre_contiene.lower() in os.path.basename(f).lower()]
    if not hits:
        return None
    return openpyxl.load_workbook(hits[0], data_only=True)

# ----------------------------------------------------------------------------
# 1) Dimensión EMPRESA a partir del export de Mapon (BASE DE DATOS JUNIO)
# ----------------------------------------------------------------------------
def cargar_dimension_empresa():
    """Construye:
       vin2cid:   VIN -> ID Cliente (Mapon)
       cid_info:  ID Cliente -> {nombre, manager, creado, total_unidades, online, offline, cuenta}
       cid_cont:  ID Cliente -> [ {nombre, tipo, email, telefono} ... ]
       name2cid:  nombre normalizado -> ID Cliente (fallback)
    """
    vin2cid, cid_info, cid_cont, name2cid = {}, {}, defaultdict(list), {}
    wb = wb_of("BASE_DE_DATOS")
    if not wb:
        return vin2cid, cid_info, cid_cont, name2cid

    # Detalle_Clientes -> info de empresa
    if "Copia de Detalle_Clientes" in wb.sheetnames:
        ws = wb["Copia de Detalle_Clientes"]
        for r in ws.iter_rows(min_row=2, values_only=True):
            cid = clean(r[0]).replace(".0", "")
            if not cid:
                continue
            nombre = clean(r[2])
            cid_info[cid] = {
                "nombre": nombre,
                "cuenta": clean(r[1]),
                "manager": clean(r[4]),
                "creado": parse_date(r[5]),
                "total_unidades": clean(r[6]).replace(".0", ""),
                "online": clean(r[7]).replace(".0", ""),
                "offline": clean(r[8]).replace(".0", ""),
            }
            if nombre:
                name2cid.setdefault(fold(nombre), cid)

    # Base de datos -> VIN -> ID Cliente (+ nombre por si falta en Detalle)
    if "Base de datos" in wb.sheetnames:
        ws = wb["Base de datos"]
        hdr_i, hdr = find_header(ws)
        vin_i = hdr.get("vin", 4)
        for r in ws.iter_rows(min_row=hdr_i + 2, values_only=True):
            vin = norm_vin(r[vin_i]) if vin_i < len(r) else None
            if not is_vin(vin):
                continue
            cid = clean(r[7]).replace(".0", "") if len(r) > 7 else ""
            if cid:
                vin2cid.setdefault(vin, cid)
                if cid not in cid_info and len(r) > 0:
                    cid_info[cid] = {"nombre": clean(r[0]), "cuenta": clean(r[17]) if len(r) > 17 else "",
                                     "manager": "", "creado": None, "total_unidades": "", "online": "", "offline": ""}

    # Detalle_Unidades (dos variantes) -> VIN -> ID Cliente
    for sh, vcol, ccol in [("Copia de Detalle_Unidades", 1, 2),
                            ("Copia de Detalle_Unidades 1", 2, 3)]:
        if sh in wb.sheetnames:
            for r in wb[sh].iter_rows(min_row=2, values_only=True):
                if len(r) <= max(vcol, ccol):
                    continue
                vin = norm_vin(r[vcol])
                cid = clean(r[ccol]).replace(".0", "")
                if is_vin(vin) and cid:
                    vin2cid.setdefault(vin, cid)

    # Detalle_Usuarios -> contactos por empresa
    if "Copia de Detalle_Usuarios" in wb.sheetnames:
        for r in wb["Copia de Detalle_Usuarios"].iter_rows(min_row=2, values_only=True):
            if len(r) < 7:
                continue
            cid = clean(r[2]).replace(".0", "")
            if not cid:
                continue
            cid_cont[cid].append({
                "nombre": clean(r[3]), "tipo": fold(r[4]),
                "email": clean(r[5]), "telefono": clean(r[6]), "fuente": "Detalle_Usuarios",
            })

    # Copia de Advance -> contactos por NOMBRE de empresa
    if "Copia de Advance" in wb.sheetnames:
        for r in wb["Copia de Advance"].iter_rows(min_row=2, values_only=True):
            if len(r) < 5:
                continue
            nom = fold(r[0])
            cid = name2cid.get(nom)
            if not cid:
                continue
            cid_cont[cid].append({
                "nombre": clean(r[1]), "tipo": fold(r[2]),
                "email": clean(r[3]), "telefono": clean(r[4]), "fuente": "Advance",
            })
    wb.close()
    return vin2cid, cid_info, cid_cont, name2cid

def _fecha_en_texto(seg):
    """Primera fecha dd/mm/yyyy, dd.mm.yyyy o yyyy-mm-dd dentro de un texto."""
    m = re.search(r"(\d{4})[-/.](\d{1,2})[-/.](\d{1,2})", seg)
    if m:
        y, mo, da = m.groups()
    else:
        m = re.search(r"(\d{1,2})[-/.](\d{1,2})[-/.](\d{2,4})", seg)
        if not m:
            return None
        da, mo, y = m.groups()
        if len(y) == 2:
            y = "20" + y
    try:
        d = dt.date(int(y), int(mo), int(da))
        return d if 2019 <= d.year <= 2035 else None
    except ValueError:
        return None

def parse_notes(notes):
    """De las 'Notas de Referencia' del CRM saca (renovacion, activacion).
    Ej.: 'Activación: 26.02.2026 / Próxima Renovación: 26/02/2027'."""
    if not notes:
        return None, None
    renov = activacion = None
    for linea in re.split(r"[\n\r;]+", str(notes)):
        l = fold(linea)
        if "renov" in l:
            renov = renov or _fecha_en_texto(linea)
        if "activ" in l or "alta" in l:
            activacion = activacion or _fecha_en_texto(linea)
    # si no hubo etiqueta de renovación pero hay UNA sola fecha futura, no asumimos
    return renov, activacion

def cargar_crm(crm_dir):
    """Lee el dump del CRM en vivo (companies.json, cars.json).
    Devuelve:
      crm_vin:  VIN -> {alta, notes_renov, companyId, last_data, make, model, notes}
      crm_comp: companyId -> {name, statusTill, creado, manager, email, segment}
      crm_units: companyId -> nº de cars en el CRM
    """
    import json
    crm_vin, crm_comp, crm_units = {}, {}, defaultdict(int)
    cp = os.path.join(crm_dir, "companies.json")
    cr = os.path.join(crm_dir, "cars.json")
    if not (os.path.exists(cp) and os.path.exists(cr)):
        return crm_vin, crm_comp, crm_units
    for c in json.load(open(cp, encoding="utf-8")):
        cid = str(c.get("id"))
        crm_comp[cid] = {
            "name": clean(c.get("name")),
            "statusTill": clean(c.get("statusTill")),
            "creado": parse_date((c.get("createdAt") or {}).get("raw") if isinstance(c.get("createdAt"), dict) else c.get("createdAt")),
            "manager": clean(((c.get("manager") or {}).get("name", "") + " " + (c.get("manager") or {}).get("surname", "")).strip()),
            "email": clean(c.get("email")),
            "segment": clean((c.get("segment") or {}).get("name") if isinstance(c.get("segment"), dict) else ""),
        }
    for car in json.load(open(cr, encoding="utf-8")):
        vin = norm_vin(car.get("vinNumber"))
        cid = str(car.get("companyId"))
        if cid:
            crm_units[cid] += 1
        if not is_vin(vin):
            continue
        renov, activ = parse_notes(car.get("notes"))
        crm_vin[vin] = {
            "alta": parse_date(car.get("createDateTime")) or activ,
            "notes_renov": renov,
            "companyId": cid,
            "last_data": parse_date(car.get("lastDataReceived")),
            "make": clean(car.get("make")), "model": clean(car.get("model")),
            "notes": clean(car.get("notes")),
        }
    return crm_vin, crm_comp, crm_units

def cargar_telemetria():
    """VIN -> {ultima_conexion: date, estado: 'Online'/'Offline', vigencia: str}
    desde el export de Mapon (BASE DE DATOS). Sirve para decidir si una unidad
    sigue activa antes de proyectarle una fecha de renovación estimada."""
    tel = {}
    wb = wb_of("BASE_DE_DATOS")
    if not wb:
        return tel
    for sh in ("Base de datos", "Hoja 5"):
        if sh not in wb.sheetnames:
            continue
        ws = wb[sh]
        hdr_i, hdr = find_header(ws)
        vin_i = hdr.get("vin", 4)
        for r in ws.iter_rows(min_row=hdr_i + 2, values_only=True):
            if vin_i >= len(r):
                continue
            vin = norm_vin(r[vin_i])
            if not is_vin(vin):
                continue
            # columnas fijas del export: 12=Vigencia, 15=Última Conexión, 16=Estado
            vig = clean(r[12]) if len(r) > 12 else ""
            uconx = parse_date(r[15]) if len(r) > 15 else None
            estado = clean(r[16]) if len(r) > 16 else ""
            cur = tel.get(vin)
            if cur is None or (uconx and (cur["ultima_conexion"] is None or uconx > cur["ultima_conexion"])):
                tel[vin] = {"ultima_conexion": uconx, "estado": estado, "vigencia": vig}
    wb.close()
    return tel

def add_years(d, n):
    """d + n años, seguro para 29-feb."""
    try:
        return d.replace(year=d.year + n)
    except ValueError:
        return d.replace(year=d.year + n, day=28)

def proxima_renovacion(anchor, hoy):
    """Siguiente aniversario anual del 'anchor' que sea >= hoy.
    Si el anchor ya es futuro, se respeta tal cual."""
    if not anchor:
        return None
    if anchor >= hoy:
        return anchor
    k = max(0, hoy.year - anchor.year)
    cand = add_years(anchor, k)
    while cand < hoy:
        k += 1
        cand = add_years(anchor, k)
    return cand

TIPO_PREF = {"admin": 0, "user_all": 1, "user": 2, "": 3}

def mejor_contacto(cands):
    """Elige el contacto más completo (con tel y correo), prefiriendo admin/user_all."""
    if not cands:
        return {"nombre": "", "telefono": "", "correo": ""}
    def score(c):
        completo = (1 if c.get("telefono") else 0) + (1 if c.get("email") else 0)
        return (-completo, TIPO_PREF.get(c.get("tipo", ""), 3))
    c = sorted(cands, key=score)[0]
    return {"nombre": c.get("nombre", ""), "telefono": c.get("telefono", ""), "correo": c.get("email", "")}

# ----------------------------------------------------------------------------
# 2) Recolección de registros por VIN desde TODAS las hojas relevantes
# ----------------------------------------------------------------------------
def prioridad_renov(archivo, hoja):
    a = archivo.lower()
    h = fold(hoja)
    if "actualizado" in a:
        return 1   # Operaciones (validada)
    if "230126_eri" in a or "230126_eri_ok" in a or "renovaciones_230126" in a:
        if h in ("unidades nuevas", "hoja 34", "renovaciones missa oct", "agosto"):
            return 2
        return None  # 230126 device = sólo alta
    if "base_de_datos" in a:
        return 3   # export API
    if "bot_api" in a:
        return 4   # API cruda
    return 9

def recolectar():
    registros = []  # cada uno: dict
    for f in sorted(glob.glob(os.path.join(SRC, "*.xlsx"))):
        base = os.path.basename(f)
        wb = openpyxl.load_workbook(f, data_only=True)
        for ws in wb.worksheets:
            if ws.max_row is None or ws.max_row < 2:
                continue
            hdr_i, hdr = find_header(ws)
            if hdr_i < 0 or "vin" not in hdr:
                continue
            pr = prioridad_renov(base, ws.title)
            vin_i = hdr["vin"]
            for r in ws.iter_rows(min_row=hdr_i + 2, values_only=True):
                if vin_i >= len(r):
                    continue
                vin = norm_vin(r[vin_i])
                if not is_vin(vin):
                    continue
                def g(key):
                    i = hdr.get(key)
                    return r[i] if (i is not None and i < len(r)) else None
                renov = parse_date(g("fecha_renov"))
                alta = parse_date(g("fecha_alta"))
                cont = clean(g("contacto_nombre")) or clean(g("contacto_alt"))
                disp = norm_dispositivo(clean(g("modelo_gps"))) or norm_dispositivo(clean(g("dispositivo")))
                registros.append({
                    "vin": vin,
                    "empresa": clean(g("empresa")),
                    "dispositivo": disp,
                    "fecha_renov": renov,
                    "fecha_alta": alta,
                    "contacto": cont,
                    "telefono": clean(g("telefono")),
                    "correo": clean(g("correo")),
                    "unidades_raw": clean(g("unidades")).replace(".0", ""),
                    "clasif_raw": clean(g("clasificacion")),
                    "estatus": clean(g("estatus")),
                    "vigencia": clean(g("vigencia")),
                    "archivo": base,
                    "hoja": ws.title,
                    "prioridad": pr,
                })
        wb.close()
    return registros

# ----------------------------------------------------------------------------
# 3) Consolidación por VIN
# ----------------------------------------------------------------------------
def tipo_cliente(n):
    if not n:
        return ""
    if n >= 20:
        return "AAA"
    if n >= 10:
        return "AA"
    if n >= 1:
        return "A"
    return ""

def consolidar(registros, dim, tel=None, crm=None):
    vin2cid, cid_info, cid_cont, name2cid = dim
    tel = tel or {}
    crm_vin, crm_comp, crm_units = crm or ({}, {}, {})
    hoy = dt.date.today()
    por_vin = defaultdict(list)
    for x in registros:
        por_vin[x["vin"]].append(x)

    # Conteo de unidades reales por empresa (VINs distintos observados)
    vins_por_cid = defaultdict(set)
    vins_por_nombre = defaultdict(set)
    for vin, regs in por_vin.items():
        cid = vin2cid.get(vin)
        if cid:
            vins_por_cid[cid].add(vin)
        else:
            # agrupar por nombre normalizado del primer registro con empresa
            nom = next((fold(r["empresa"]) for r in regs if r["empresa"]), "")
            if nom:
                vins_por_nombre[nom].add(vin)

    unidades = []   # filas finales por VIN
    conflictos = []
    for vin, regs in sorted(por_vin.items()):
        cv = crm_vin.get(vin)               # datos vivos del CRM para este VIN
        cid = (cv["companyId"] if cv and cv.get("companyId") else None) or vin2cid.get(vin)

        # ---- FECHA DE RENOVACIÓN: por prioridad, luego marcamos conflicto ----
        # P0 = "Próxima Renovación" escrita por los analistas en las Notas del CRM (verdad final).
        con_fecha = [r for r in regs if r["fecha_renov"] and r["prioridad"] is not None]
        if cv and cv.get("notes_renov"):
            con_fecha.insert(0, {"fecha_renov": cv["notes_renov"], "prioridad": 0,
                                 "archivo": "CRM en vivo", "hoja": "Notas de Referencia"})
        elegida = None
        elegida_src = ""
        if con_fecha:
            con_fecha.sort(key=lambda r: (r["prioridad"], -r["fecha_renov"].toordinal()))
            best = con_fecha[0]
            elegida = best["fecha_renov"]
            elegida_src = f'{best["archivo"]}::{best["hoja"]}'
            # ¿discrepan las fuentes?
            fechas = {r["fecha_renov"] for r in con_fecha}
            if len(fechas) > 1:
                spread = (max(fechas) - min(fechas)).days
                if spread > TOLERANCIA_DIAS:
                    detalle = sorted({(r["fecha_renov"].isoformat(), r["prioridad"],
                                       f'{r["archivo"]}::{r["hoja"]}') for r in con_fecha})
                    conflictos.append({
                        "vin": vin, "elegida": elegida.isoformat(), "fuente_elegida": elegida_src,
                        "dias_diferencia": spread,
                        "fechas": " | ".join(f"{d} (P{p}) {s}" for d, p, s in detalle),
                    })

        # ---- FECHA DE ALTA: preferimos Unidades nuevas > Fleet/instalación > 230126 ----
        altas = [r for r in regs if r["fecha_alta"]]
        def alta_pri(r):
            h = fold(r["hoja"])
            if "unidades nuevas" in h:
                return 0
            if "fleet export" in h:
                return 1
            return 2
        alta = None
        if altas:
            altas.sort(key=alta_pri)
            alta = altas[0]["fecha_alta"]
        # El CRM (cars.createDateTime) es la fecha de alta AUTORITATIVA.
        if cv and cv.get("alta"):
            alta = cv["alta"]

        # ---- EMPRESA (el CRM en vivo manda; luego export; luego archivos) ----
        if cv and cid in crm_comp and crm_comp[cid].get("name"):
            empresa = crm_comp[cid]["name"]
        elif cid and cid in cid_info and cid_info[cid]["nombre"]:
            empresa = cid_info[cid]["nombre"]
        else:
            empresa = next((r["empresa"] for r in regs if r["empresa"]), "")

        # ---- UNIDADES + TIPO ----
        n_units = None
        if cid and cid in crm_units and crm_units[cid]:
            n_units = crm_units[cid]            # conteo vivo de vehículos en el CRM
        elif cid and cid in cid_info and cid_info[cid].get("total_unidades"):
            try:
                n_units = int(float(cid_info[cid]["total_unidades"]))
            except ValueError:
                n_units = None
        if not n_units:
            if cid:
                n_units = len(vins_por_cid.get(cid, ())) or None
            else:
                n_units = len(vins_por_nombre.get(fold(empresa), ())) or None
        tipo = tipo_cliente(n_units)

        # ---- CONTACTO ----
        # 1) contacto operativo a nivel VIN (el más específico)
        cont_vin = next((r for r in regs if r["contacto"] or r["correo"] or r["telefono"]), None)
        nombre_c = telefono_c = correo_c = ""
        if cont_vin and (cont_vin["contacto"] or cont_vin["correo"] or cont_vin["telefono"]):
            nombre_c, telefono_c, correo_c = cont_vin["contacto"], cont_vin["telefono"], cont_vin["correo"]
        # 2) completar con el mejor contacto de la empresa (CRM export)
        if cid and (not telefono_c or not correo_c or not nombre_c):
            mc = mejor_contacto(cid_cont.get(cid, []))
            nombre_c = nombre_c or mc["nombre"]
            telefono_c = telefono_c or mc["telefono"]
            correo_c = correo_c or mc["correo"]

        estatus = next((r["estatus"] for r in regs if r["estatus"]), "")
        dispositivo = next((r["dispositivo"] for r in regs if r.get("dispositivo")), "")
        t = tel.get(vin, {})
        vigencia = t.get("vigencia") or next((r["vigencia"] for r in regs if r["vigencia"]), "")
        uconx = t.get("ultima_conexion")
        estado = t.get("estado", "")
        # El CRM trae la última recepción más fresca: manda si es más reciente.
        if cv and cv.get("last_data") and (uconx is None or cv["last_data"] > uconx):
            uconx = cv["last_data"]
            if not estado:
                estado = "Online" if (hoy - cv["last_data"]).days <= 30 else "Offline"

        # ---- ¿unidad activa? (para decidir si proyectamos renovación) ----
        activa = None  # True/False/None(desconocido)
        if fold(vigencia) == "baja" or fold(estatus) in ("no renovo", "baja"):
            activa = False
        elif fold(estado) == "online" or fold(vigencia) in ("activo", "vigente"):
            activa = True
        elif uconx is not None:
            activa = (hoy - uconx).days <= 120
        elif fold(estado) == "offline":
            activa = False

        # ---- SITUACIÓN: Activa / Baja / Pendiente de revisión ----
        # Una unidad dada de baja PERO que sigue reportando GPS no es una baja
        # limpia: queda "Pendiente de revisión" para que Operaciones decida.
        reporta = (fold(estado) == "online") or (uconx is not None and (hoy - uconx).days <= 30)
        if activa is False and reporta:
            situacion = "Pendiente de revisión"
        elif activa is False:
            situacion = "Baja"
        else:
            situacion = "Activa"

        # ---- PRÓXIMA RENOVACIÓN (accionable, siempre >= hoy) ----
        # Ancla por orden de confianza: fecha real > alta > última conexión > creación empresa.
        if elegida:
            ancla, base = elegida, "archivo"
        elif alta:
            ancla, base = alta, "estimada: aniversario de alta"
        elif uconx:
            ancla, base = uconx, "estimada: ancla última conexión"
        elif cid and cid in cid_info and cid_info[cid].get("creado"):
            ancla, base = cid_info[cid]["creado"], "estimada: ancla creación de empresa"
        else:
            ancla, base = None, "sin dato suficiente"

        prox = proxima_renovacion(ancla, hoy) if ancla else None
        # Regla del cliente: a las unidades SIN fecha real se les dan 90 días
        # adicionales (12 meses de servicio + prórroga por incertidumbre de datos).
        if prox and base.startswith("estimada"):
            prox = prox + dt.timedelta(days=90)
        proxima = prox.isoformat() if prox else ""
        # Si la unidad NO está activa, no inventamos renovación: es recuperación, no renovación.
        if base != "archivo" and activa is False:
            origen = "no proyectada (" + situacion.lower() + ")"
            proxima = ""
        elif base == "archivo" and prox and elegida and prox != elegida:
            origen = "archivo (rodada a próximo aniversario)"
        elif base.startswith("estimada"):
            origen = base + " +90d"
        else:
            origen = base
        estimada = proxima if base.startswith("estimada") else ""
        final = proxima
        # Fecha límite de cobranza: 90 días después del vencimiento.
        limite_gracia = ""
        if prox:
            limite_gracia = (prox + dt.timedelta(days=90)).isoformat()

        unidades.append({
            "vin": vin,
            "fecha_alta": alta.isoformat() if alta else "",
            "fecha_renovacion": elegida.isoformat() if elegida else "",
            "fuente_fecha_renovacion": elegida_src,
            "renovacion_estimada": estimada,
            "fecha_renovacion_final": final,
            "limite_gracia": limite_gracia,
            "origen_fecha": origen,
            "empresa": empresa,
            "id_cliente_mapon": cid or "",
            "contacto": nombre_c,
            "telefono": telefono_c,
            "correo": correo_c,
            "numero_unidades": n_units or "",
            "tipo_cliente": tipo,
            "dispositivo": dispositivo,
            "estatus_renovacion": estatus,
            "vigencia": vigencia,
            "ultima_conexion": uconx.isoformat() if uconx else "",
            "estado": estado,
            "activa": {True: "Sí", False: "No", None: "?"}[activa],
            "situacion": situacion,
            "fuentes_vistas": len(regs),
        })
    return unidades, conflictos, (vins_por_cid, vins_por_nombre)

# ----------------------------------------------------------------------------
# 4) Salidas
# ----------------------------------------------------------------------------
COLS_UNIDADES = [
    ("vin", "VIN"),
    ("fecha_alta", "FECHA DE ALTA"),
    ("fecha_renovacion", "FECHA DE RENOVACIÓN"),
    ("fuente_fecha_renovacion", "FUENTE FECHA RENOVACIÓN"),
    ("fecha_renovacion_final", "PRÓXIMA RENOVACIÓN"),
    ("limite_gracia", "LÍMITE PRÓRROGA (+90d)"),
    ("origen_fecha", "ORIGEN PRÓXIMA RENOVACIÓN"),
    ("empresa", "NOMBRE EMPRESA"),
    ("id_cliente_mapon", "ID CLIENTE (MAPON)"),
    ("contacto", "NOMBRE CONTACTO"),
    ("telefono", "TELÉFONO"),
    ("correo", "CORREO"),
    ("numero_unidades", "NÚMERO DE UNIDADES"),
    ("tipo_cliente", "TIPO DE CLIENTE"),
    ("dispositivo", "TIPO DE DISPOSITIVO"),
    ("estatus_renovacion", "ESTATUS RENOVACIÓN"),
    ("vigencia", "VIGENCIA"),
    ("ultima_conexion", "ÚLTIMA CONEXIÓN"),
    ("estado", "ESTADO"),
    ("activa", "ACTIVA"),
    ("situacion", "SITUACIÓN"),
    ("fuentes_vistas", "# FUENTES"),
]

def construir_empresas(unidades):
    emp = {}
    for u in unidades:
        key = u["id_cliente_mapon"] or fold(u["empresa"]) or u["vin"]
        e = emp.setdefault(key, {
            "empresa": u["empresa"], "id_cliente_mapon": u["id_cliente_mapon"],
            "contacto": u["contacto"], "telefono": u["telefono"], "correo": u["correo"],
            "numero_unidades": u["numero_unidades"], "tipo_cliente": u["tipo_cliente"],
            "vins": 0, "con_fecha": 0, "prox_renovacion": "",
        })
        e["vins"] += 1
        if u["fecha_renovacion"]:
            e["con_fecha"] += 1
            if not e["prox_renovacion"] or u["fecha_renovacion"] < e["prox_renovacion"]:
                e["prox_renovacion"] = u["fecha_renovacion"]
        # completar contacto si faltaba
        for k in ("contacto", "telefono", "correo"):
            if not e[k] and u[k]:
                e[k] = u[k]
    return list(emp.values())

def escribir(salida, unidades, conflictos, empresas, stats):
    os.makedirs(salida, exist_ok=True)
    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = "Unidades"
    ws.append([h for _, h in COLS_UNIDADES])
    for u in unidades:
        ws.append([u[k] for k, _ in COLS_UNIDADES])

    we = wb.create_sheet("Empresas")
    ecols = ["NOMBRE EMPRESA", "ID CLIENTE (MAPON)", "NÚMERO DE UNIDADES", "TIPO DE CLIENTE",
             "VINS EN BASE", "VINS CON FECHA RENOV", "PRÓXIMA RENOVACIÓN",
             "NOMBRE CONTACTO", "TELÉFONO", "CORREO"]
    we.append(ecols)
    for e in sorted(empresas, key=lambda x: (-(int(x["numero_unidades"]) if str(x["numero_unidades"]).isdigit() else 0))):
        we.append([e["empresa"], e["id_cliente_mapon"], e["numero_unidades"], e["tipo_cliente"],
                   e["vins"], e["con_fecha"], e["prox_renovacion"],
                   e["contacto"], e["telefono"], e["correo"]])

    wc = wb.create_sheet("Conflictos")
    wc.append(["VIN", "FECHA ELEGIDA", "FUENTE ELEGIDA", "DÍAS DE DIFERENCIA", "TODAS LAS FECHAS (fecha (Pprioridad) fuente)"])
    for c in sorted(conflictos, key=lambda x: -x["dias_diferencia"]):
        wc.append([c["vin"], c["elegida"], c["fuente_elegida"], c["dias_diferencia"], c["fechas"]])

    wm = wb.create_sheet("Metodologia")
    for ln in stats:
        wm.append([ln])

    out_xlsx = os.path.join(salida, "base_homologada.xlsx")
    wb.save(out_xlsx)

    # CSV por VIN (encabezados alineados al CRM de Mapon)
    with open(os.path.join(salida, "clientes_unidades.csv"), "w", newline="", encoding="utf-8-sig") as fh:
        w = csv.writer(fh)
        w.writerow([h for _, h in COLS_UNIDADES])
        for u in unidades:
            w.writerow([u[k] for k, _ in COLS_UNIDADES])

    with open(os.path.join(salida, "empresas.csv"), "w", newline="", encoding="utf-8-sig") as fh:
        w = csv.writer(fh)
        w.writerow(ecols)
        for e in empresas:
            w.writerow([e["empresa"], e["id_cliente_mapon"], e["numero_unidades"], e["tipo_cliente"],
                        e["vins"], e["con_fecha"], e["prox_renovacion"],
                        e["contacto"], e["telefono"], e["correo"]])
    return out_xlsx

# ----------------------------------------------------------------------------
def main():
    global SRC
    ap = argparse.ArgumentParser()
    ap.add_argument("--fuentes", default=os.path.join(os.path.dirname(__file__), "fuentes"))
    ap.add_argument("--salida", default=os.path.join(os.path.dirname(__file__), "salida"))
    ap.add_argument("--crm", default=os.path.join(os.path.dirname(__file__), "salida", "_crm_raw"),
                    help="carpeta con el dump del CRM (companies.json, cars.json)")
    args = ap.parse_args()
    SRC = args.fuentes

    print("· Cargando dimensión de empresa (export Mapon)…")
    dim = cargar_dimension_empresa()
    print(f"  VIN→IDcliente: {len(dim[0])}  ·  empresas: {len(dim[1])}  ·  empresas c/contacto: {len(dim[2])}")

    print("· Cargando CRM en vivo (si hay dump)…")
    crm = cargar_crm(args.crm)
    crm_renov = sum(1 for v in crm[0].values() if v.get("notes_renov"))
    print(f"  VINs en CRM: {len(crm[0])}  ·  empresas CRM: {len(crm[1])}  ·  renovación en notas: {crm_renov}")

    print("· Cargando telemetría por VIN (última conexión / estado)…")
    tel = cargar_telemetria()
    print(f"  VINs con telemetría: {len(tel)}")

    print("· Recolectando registros por VIN de todas las hojas…")
    registros = recolectar()
    print(f"  registros crudos: {len(registros)}")

    print("· Consolidando por VIN…")
    unidades, conflictos, _ = consolidar(registros, dim, tel, crm)
    empresas = construir_empresas(unidades)

    # Estadísticas / metodología
    total = len(unidades)
    con_renov = sum(1 for u in unidades if u["fecha_renovacion"])
    con_alta = sum(1 for u in unidades if u["fecha_alta"])
    con_contacto = sum(1 for u in unidades if u["telefono"] or u["correo"])
    con_final = sum(1 for u in unidades if u["fecha_renovacion_final"])
    estimadas = sum(1 for u in unidades if u["renovacion_estimada"])
    sin_dato = sum(1 for u in unidades if not u["fecha_renovacion_final"])
    por_origen = defaultdict(int)
    for u in unidades:
        por_origen[u["origen_fecha"]] += 1
    por_situacion = defaultdict(int)
    for u in unidades:
        por_situacion[u["situacion"]] += 1
    por_tipo = defaultdict(int)
    for e in empresas:
        por_tipo[e["tipo_cliente"] or "(sin clasificar)"] += 1
    hoy = dt.date.today()
    stats = [
        "METODOLOGÍA DE HOMOLOGACIÓN — base de clientes / renovaciones Advance",
        "",
        f"Generado: {hoy.isoformat()}",
        f"Tolerancia de conflicto entre fuentes: {TOLERANCIA_DIAS} días",
        "",
        "Prioridad de FECHA DE RENOVACIÓN cuando hay fechas distintas:",
        "  P1) RENOVACIONES ACTUALIZADO (Operaciones, validada a mano)",
        "  P2) Operaciones complementarias (Unidades nuevas / Hoja 34 / Missa / Agosto)",
        "  P3) Export BASE DE DATOS JUNIO (snapshot de la API)",
        "  P4) Bot_API (jalado crudo de la API)",
        "  El CRM en vivo es la verdad final (reconciliar_crm.py).",
        "",
        "Clasificación de cliente por número de unidades:",
        "  AAA = 20+   ·  AA = 10–19   ·  A = 1–9",
        "",
        "RESULTADOS",
        f"  Unidades (VIN) homologadas: {total}",
        f"  Con fecha de renovación REAL (de archivos): {con_renov}  ({100*con_renov//max(total,1)}%)",
        f"  Con fecha de renovación ESTIMADA: {estimadas}",
        f"  Con FECHA RENOVACIÓN FINAL (real+estimada): {con_final}  ({100*con_final//max(total,1)}%)",
        f"  Sin fecha (ni real ni estimable): {sin_dato}",
        f"  Con fecha de alta: {con_alta}",
        f"  Con teléfono o correo: {con_contacto}",
        f"  Empresas distintas: {len(empresas)}",
        f"  Conflictos de fecha (> {TOLERANCIA_DIAS} días): {len(conflictos)}",
        "",
        "Situación de las unidades:",
    ] + [f"  {k}: {v}" for k, v in sorted(por_situacion.items(), key=lambda x: -x[1])] + [
        "",
        "Origen de la fecha de renovación:",
    ] + [f"  {k}: {v}" for k, v in sorted(por_origen.items(), key=lambda x: -x[1])] + [
        "",
        "Empresas por tipo:",
    ] + [f"  {k}: {v}" for k, v in sorted(por_tipo.items())]

    out = escribir(args.salida, unidades, conflictos, empresas, stats)
    print("\n".join(stats))
    print(f"\n✓ Escrito: {out}")
    print(f"✓ CSVs en: {args.salida}")


if __name__ == "__main__":
    main()
